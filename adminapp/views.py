from django.contrib.auth.decorators import login_required
from django.http.response import Http404, HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from account.models import User, Profile
from adminapp import forms
from adminapp.models import District, Municipality, Panchayath, Corporation, VTCListExcel
# , VTCProfileForm
from adminapp.forms import AddVTCAdminForm, UploadFileForm
from django.core.mail import BadHeaderError, message, send_mail
from django.contrib.auth.decorators import login_required
from account.views import home_view
from django.conf import settings
import random
from django.contrib import messages
from vtc_exact_app.models import VtcModel

# from courses.models import Category, Course
from courses.models import Enrollment
from courses.forms import add_student_form_enroll, add_student_form_profile, add_student_form_user

# Sudden
from django.http import HttpResponseBadRequest
from django.template import RequestContext
from openpyxl import load_workbook
from django.db import IntegrityError


# @login_required
# def add_new_vtc(request):
#     if request.user.user_type == 3:
#         vtc_user_form = AddVTCAdminForm(None)
#         vtc_profile_form = VTCProfileForm(None)
#         print("usertype is rutronix")
#         if request.method == 'POST':
#             print("going to check form is valid")
#             vtc_user_form = AddVTCAdminForm(request.POST)
#             vtc_profile_form = VTCProfileForm(request.POST)
#             if vtc_user_form.is_valid() and vtc_profile_form.is_valid():
#                 name = vtc_user_form.cleaned_data.get('name')
#                 vtc_name = vtc_profile_form.cleaned_data.get('vtc_name')
#                 email = vtc_user_form.cleaned_data.get('email')
#                 print(email)
#                 contact_number = vtc_user_form.cleaned_data.get(
#                     'contact_number')
#                 address = vtc_profile_form.cleaned_data.get('address')
#                 district = vtc_profile_form.cleaned_data.get('district')
#                 municipality = vtc_profile_form.cleaned_data.get(
#                     'municipality')
#                 code = vtc_profile_form.cleaned_data.get('code')
#                 a, b, c, d = True if request.POST.get('psc') else False, True if request.POST.get(
#                     'ssc') else False, True if request.POST.get('upsc') else False, True if request.POST.get('bank_exam') else False
#                 password = User.objects.make_random_password(
#                     length=14, allowed_chars="abcdefghjkmnpqrstuvwxyz01234567889!@#$%^&*()")
#                 if not User.objects.filter(email=email):
#                     # if district == "Select District":
#                     # vtc_profile_form.add_error(
#                     #    field='district', error="Choose District")
#                     # else:
#                     user = User.objects.create(
#                         username=email, name=name, contact_number=contact_number, email=email, user_type=2)
#                     user.set_password(password)
#                     user.save()
#                     Profile.objects.create(user=user, vtc_name=vtc_name, address=address, district=district,
#                                            municipality=municipality, code=code, psc=a, ssc=b, upsc=c, bank_exam=d)
#                     #send_mail('Credentials - Vijayaveedhi',f'Greetings,\n\nWelcome to Rutronix Vijayaveedhi. These are your login credentials.\nsite: vijayaveedhi.org \nLogin Email: {email}, \nPassword: {password}.\n\n\n Rutronix Vijayaveedhi Team','ragikrishna9610@gmail.com', [email], fail_silently=False)
#                     subject = 'welcome to VIJAYAVEEDHI'
#                     message = f'Hi your login email id is {email},your password is {password} thank you for registering in vijayaveedhi.'
#                     email_from = settings.EMAIL_HOST_USER
#                     recipient_list = [user.email, ]
#                     send_mail(subject, message, email_from, recipient_list)
#                     return redirect('home')
#                 else:
#                     vtc_user_form.add_error(
#                         error="User already exists", field="email")
#                 # send_mail('Credentials - IT Kerala',
#                 #           f'Greetings,\n\nThese are your login credentials for IT Kerala. \nLogin Email: {email}, \nPassword: {password}.\n\n\n IT Kerala Team',
#                 #           'noreply@itkeralaedu.com', [email], fail_silently=False)
#                 return redirect('vtc')
#             else:
#                 print('error')
#         context = {
#             'vtc_user_form': vtc_user_form,
#             'vtc_profile_form': vtc_profile_form,
#         }
#         return render(request, 'add-vtc.html', context)
#     else:
#         print("eroor occur")


def load_municipalities(request):
    district_id = request.GET.get('district_id')
    municipalities = Municipality.objects.filter(district_id=district_id)
    return render(request, 'municipality_dropdown.html', {'municipalities': municipalities})
    # return JsonResponse(list(cities.values('id', 'name')), safe=False)


def view_vtcs(request):
    # data = Profile.objects.filter(user_id__user_type=2)
    data = VtcModel.objects.filter(user_id__user_type=2)
    return render(request, 'vtc-list.html', {'datas': data})


@login_required
def view_student_vtc(request):
    if request.user.user_type == 2:
        
        data = Enrollment.objects.filter(
            user_id__user_type=1, study_center=get_object_or_404(VtcModel, user=request.user))
        return render(request, 'vtc-student-list.html', {'datas': data})
    else:
        raise Http404


@login_required
def view_student_admin(request, code):
    if request.user.user_type == 3 or request.user.user_type == 4:
        data = Enrollment.objects.filter(
            user_id__user_type=1, study_center__code=code)
        return render(request, 'vtc-student-list.html', {'datas': data})
    else:
        raise Http404



# def update(request, id):
#     employee = Employee.objects.get(id=id)
#     form = EmployeeForm(request.POST, instance = employee)
#     if form.is_valid():
#         form.save()
#         return redirect("/show")
#     return render(request, 'edit.html', {'employee': employee})


def destroy(request, id):
    vtc_item = User.objects.get(id=id)
    vtc_item.delete()
    return redirect('home')


def add_student_demo(request):
    return render(request, 'add-student.html')


@login_required
def Enrollment_details(request):
    student_user_form = add_student_form_user(None)
    student_profile_form = add_student_form_profile(None)
    student_enrollment_form = add_student_form_enroll(None)
    if request.method == 'POST':
        student_user_form = add_student_form_user(request.POST)
        student_profile_form = add_student_form_profile(request.POST)
        student_enrollment_form = add_student_form_enroll(request.POST)
        print(f'\n\nvalid data')
        print(
            f'{student_user_form.is_valid()} and {student_profile_form.is_valid()} and {student_enrollment_form.is_valid()}\n\n')
        if student_user_form.is_valid() and student_profile_form.is_valid() and student_enrollment_form.is_valid():
            username = student_user_form.cleaned_data.get('email')
            contact_number = student_user_form.cleaned_data.get('contact_number')
            name = student_user_form.cleaned_data.get('name')
            email = student_user_form.cleaned_data.get('email')
            gender = student_user_form.cleaned_data.get('gender')
            qualification = student_profile_form.cleaned_data.get('qualification')
            # code = student_profile_form.cleaned_data.get('code')
            #mode_of_training = student_enrollment_form.cleaned_data.get('mode_of_training')
            course = student_enrollment_form.cleaned_data.get('course')
            print(course)
            category = student_enrollment_form.cleaned_data.get('category')
            print(category)
            batch = student_enrollment_form.cleaned_data.get('batch')
            print(batch)
            password = User.objects.make_random_password(
                length=14, allowed_chars="abcdefghjkmnpqrstuvwxyz01234567889!@#$%^&*()")
            if not User.objects.filter(email=email):
                user = User.objects.create_user(
                    username=username, name=name, contact_number=contact_number, gender=gender, email=email)
                user.set_password(password)
                user.save()
            else:
                user = get_object_or_404(User, email=email)

            if not Profile.objects.filter(user__email=email):
                profile = Profile.objects.create(
                    user=user, qualification=qualification)
                profile.save()

            # try:
            vtc=get_object_or_404(VtcModel,user = request.user)


            enrollment = Enrollment.objects.create(
                user=user, course=course, category=category, study_center=vtc, batch=batch)

            enrollment.save()
            subject = 'welcome to VIJAYAVEEDHI'
            message = f'Hi your login email id is {email},your password is {password} thank you for registering in vijayaveedhi.'
            email_from = settings.EMAIL_HOST_USER
            recipient_list = [user.email, ]
            send_mail(subject, message, email_from, recipient_list)
            messages.success(request, "The student has been added")
            # except IntegrityError:
            #     pass
            return redirect('add_student')
        else:
            messages.error(request, "The student has already been added")
            print(f'\n\nerrors occur\n\n')
    context = {'student_user_form': student_user_form, 'student_profile_form':
               student_profile_form, 'student_enrollment_form': student_enrollment_form}
    return render(request, 'add-student-by-vtc.html', context)


def vtc_student_list(request):
    return render(request, 'vtc-student-list.html')


def upload(request):
    form = UploadFileForm(None)
    if request.method == "POST":
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            filehandle = request.FILES['excel_file']
            VTCListExcel.objects.create(excel_file=filehandle.name)
            book = load_workbook(filehandle, data_only=True)
            sheet = book.active
            max_col = sheet.max_column
            max_row = sheet.max_row
            for i in range(2, max_row + 1):
                arr = []
                # print("\n")
                print("Row ", i, " data :")
                for j in range(1, max_col + 1):
                    cell_obj = sheet.cell(row=i, column=j)
                    arr.append(cell_obj.value)
                print(arr)
                if not arr[4]:
                    print("None")
                    break
                password = User.objects.make_random_password(
                    length=14, allowed_chars="abcdefghjkmnpqrstuvwxyz01234567889!@#$%^&*()")
                try:
                    user = User.objects.create(
                        username=arr[4], name=arr[2], contact_number=arr[3], email=arr[4], user_type=2)
                    user.set_password(password)
                    user.save()
                    VtcModel.objects.create(
                        user=user, vtc_name=arr[0], code=arr[1])
                    subject = 'welcome to VIJAYAVEEDHI'
                    message = f'Hi your login email id is {arr[4]},your password is {password} thank you for registering in vijayaveedhi.'
                    email_from = settings.EMAIL_HOST_USER
                    recipient_list = [arr[4]]
                    send_mail(subject, message, email_from, recipient_list)
                except IntegrityError:
                    pass
        # return excel.make_response(filehandle.get_sheet(), "csv")
        else:
            print('\n\nInvalidform\n\n')
    return render(request, 'upload_form.html', {'form': form})
